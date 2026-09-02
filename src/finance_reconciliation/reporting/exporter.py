"""Render FinanceReportData into the operational Excel workbook.

Excel concerns only - no SQL, no Finance calculation. Every value comes
pre-computed from the marts; the exporter just lays it out and formats
it for reading.
"""

from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from finance_reconciliation.reporting.models import (
    DAILY_SUMMARY_FIELDS,
    EXCEPTION_FIELDS,
    FinanceReportData,
    FinanceReportExportResult,
)

DAILY_SHEET_NAME = "Daily Summary"
EXCEPTION_SHEET_NAME = "Exceptions"

DATE_FORMAT = "yyyy-mm-dd"
EUR_FORMAT = "#,##0.00"
COUNT_FORMAT = "#,##0"
RATE_FORMAT = "0.00%"

_DATE_FIELDS = {"business_date"}
_RATE_FIELDS = {"amount_reconciliation_rate"}
_COUNT_SUFFIXES = ("_count", "_days")


def _column_format(field: str) -> str | None:
    if field in _DATE_FIELDS:
        return DATE_FORMAT

    if field in _RATE_FIELDS:
        return RATE_FORMAT

    if field.endswith(("_eur", "_amount")):
        return EUR_FORMAT

    if field.endswith(_COUNT_SUFFIXES):
        return COUNT_FORMAT

    return None


def _write_sheet(
    worksheet: Worksheet,
    *,
    fields: tuple[str, ...],
    rows: tuple[tuple[object, ...], ...],
) -> None:
    header_font = Font(bold=True)

    for column_index, field in enumerate(
        fields,
        start=1,
    ):
        cell = worksheet.cell(
            row=1,
            column=column_index,
            value=field,
        )
        cell.font = header_font

    number_formats = [
        _column_format(field)
        for field in fields
    ]

    for row_offset, row_values in enumerate(
        rows,
        start=2,
    ):
        for column_index, value in enumerate(
            row_values,
            start=1,
        ):
            cell = worksheet.cell(
                row=row_offset,
                column=column_index,
                value=value,
            )

            number_format = number_formats[
                column_index - 1
            ]

            if (
                number_format is not None
                and value is not None
            ):
                cell.number_format = number_format

    last_column = get_column_letter(len(fields))
    last_row = len(rows) + 1

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        f"A1:{last_column}{last_row}"
    )

    for column_index, field in enumerate(
        fields,
        start=1,
    ):
        width = min(
            max(len(field) + 2, 12),
            48,
        )
        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = width


def build_workbook(
    data: FinanceReportData,
) -> Workbook:
    workbook = Workbook()

    daily_sheet = workbook.active
    daily_sheet.title = DAILY_SHEET_NAME
    _write_sheet(
        daily_sheet,
        fields=DAILY_SUMMARY_FIELDS,
        rows=tuple(
            row.as_row()
            for row in data.daily
        ),
    )

    exception_sheet = workbook.create_sheet(
        EXCEPTION_SHEET_NAME
    )
    _write_sheet(
        exception_sheet,
        fields=EXCEPTION_FIELDS,
        rows=tuple(
            row.as_row()
            for row in data.exceptions
        ),
    )

    return workbook


def export_finance_report(
    data: FinanceReportData,
    *,
    output_path: Path,
) -> FinanceReportExportResult:
    output_path = Path(output_path)
    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = build_workbook(data)

    # Atomic replace: a failed save must not leave a half-written .xlsx
    # in place of the previous good report.
    temporary_path = output_path.with_name(
        output_path.name + ".tmp"
    )
    workbook.save(temporary_path)
    os.replace(temporary_path, output_path)

    return FinanceReportExportResult(
        output_path=output_path,
        daily_row_count=len(data.daily),
        exception_row_count=len(data.exceptions),
    )
