from __future__ import annotations

from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from finance_reconciliation.reporting.exporter import (
    DAILY_SHEET_NAME,
    EXCEPTION_SHEET_NAME,
    export_finance_report,
)
from finance_reconciliation.reporting.models import (
    DAILY_SUMMARY_FIELDS,
    EXCEPTION_FIELDS,
    DailySummaryRow,
    ExceptionRow,
    FinanceReportData,
)


def _daily_row(
    *,
    business_date: date,
    product_id: str,
    rate: Decimal | None,
) -> DailySummaryRow:
    return DailySummaryRow(
        business_date=business_date,
        product_id=product_id,
        product_name=f"{product_id} name",
        product_family="CORE",
        currency="EUR",
        invoice_count=3,
        capture_count=2,
        capture_amount=Decimal("100.00"),
        valued_capture_amount_eur=Decimal("120.50"),
        unvalued_capture_count=1,
        reconciled_capture_amount_eur=Decimal("110.25"),
        open_break_capture_amount_eur=Decimal("10.25"),
        amount_reconciliation_rate=rate,
        refund_amount=Decimal("5.00"),
        chargeback_amount=Decimal("0.00"),
        exception_count=1,
        open_break_exception_count=1,
        critical_exception_count=1,
        gross_exception_amount_eur=Decimal("10.25"),
    )


def _exception_row(
    *,
    exception_code: str,
    severity: str,
) -> ExceptionRow:
    return ExceptionRow(
        exception_id=f"{exception_code}:CAPTURE:EVT-1",
        exception_code=exception_code,
        exception_status="OPEN_BREAK",
        severity=severity,
        entity_type="CAPTURE",
        entity_id="EVT-1",
        business_date=date(2026, 1, 15),
        product_id="PROD-1",
        product_name="PROD-1 name",
        currency="EUR",
        exception_amount_eur=Decimal("10.25"),
        age_days=4,
        observed_amount_eur=Decimal("100.00"),
        expected_amount_eur=Decimal("110.25"),
        difference_amount_eur=Decimal("-10.25"),
        control_source="payment_reconciliation_fact",
    )


def _report_data(
    *,
    exceptions: tuple[ExceptionRow, ...],
) -> FinanceReportData:
    return FinanceReportData(
        daily=(
            _daily_row(
                business_date=date(2026, 1, 15),
                product_id="PROD-1",
                rate=Decimal("0.912500"),
            ),
            _daily_row(
                business_date=date(2026, 1, 16),
                product_id="PROD-2",
                rate=None,
            ),
        ),
        exceptions=exceptions,
    )


def test_export_writes_both_sheets_with_headers_and_rows(tmp_path):
    output_path = (
        tmp_path / "finance_reconciliation_report.xlsx"
    )

    data = _report_data(
        exceptions=(
            _exception_row(
                exception_code="CAPTURE_AMOUNT_MISMATCH",
                severity="CRITICAL",
            ),
            _exception_row(
                exception_code="LATE_SETTLEMENT",
                severity="WARNING",
            ),
        )
    )

    result = export_finance_report(
        data,
        output_path=output_path,
    )

    assert result.output_path == output_path
    assert result.daily_row_count == 2
    assert result.exception_row_count == 2
    assert output_path.exists()
    assert output_path.stat().st_size > 0
    assert not output_path.with_name(
        output_path.name + ".tmp"
    ).exists()

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == [
        DAILY_SHEET_NAME,
        EXCEPTION_SHEET_NAME,
    ]

    daily = workbook[DAILY_SHEET_NAME]
    assert [
        cell.value for cell in daily[1]
    ] == list(DAILY_SUMMARY_FIELDS)
    assert daily.max_row == 3
    assert daily.freeze_panes == "A2"

    # Row values are passed through, not recomputed. openpyxl round-trips
    # a date as a datetime at midnight and numbers as float (Excel has
    # neither a pure-date nor a decimal type); the value is preserved.
    assert daily["A2"].value.date() == date(2026, 1, 15)
    assert daily["A2"].number_format == "yyyy-mm-dd"
    assert daily["B2"].value == "PROD-1"
    assert daily["I2"].value == 120.5
    assert daily["I2"].number_format == "#,##0.00"
    rate_cell = daily.cell(
        row=2,
        column=DAILY_SUMMARY_FIELDS.index(
            "amount_reconciliation_rate"
        )
        + 1,
    )
    assert rate_cell.value == 0.9125
    assert rate_cell.number_format == "0.00%"
    assert daily.cell(
        row=3,
        column=DAILY_SUMMARY_FIELDS.index(
            "amount_reconciliation_rate"
        )
        + 1,
    ).value is None

    exceptions = workbook[EXCEPTION_SHEET_NAME]
    assert [
        cell.value for cell in exceptions[1]
    ] == list(EXCEPTION_FIELDS)
    assert exceptions.max_row == 3
    assert exceptions["C2"].value == "OPEN_BREAK"
    assert exceptions["K2"].value == 10.25


def test_export_with_zero_exceptions_keeps_header_only_sheet(tmp_path):
    output_path = tmp_path / "report.xlsx"

    result = export_finance_report(
        _report_data(exceptions=()),
        output_path=output_path,
    )

    assert result.exception_row_count == 0

    workbook = load_workbook(output_path)
    exceptions = workbook[EXCEPTION_SHEET_NAME]

    assert [
        cell.value for cell in exceptions[1]
    ] == list(EXCEPTION_FIELDS)
    assert exceptions.max_row == 1
    assert workbook[DAILY_SHEET_NAME].max_row == 3


def test_export_overwrites_previous_report(tmp_path):
    output_path = tmp_path / "report.xlsx"

    export_finance_report(
        _report_data(
            exceptions=(
                _exception_row(
                    exception_code="A",
                    severity="CRITICAL",
                ),
            )
        ),
        output_path=output_path,
    )

    second = export_finance_report(
        _report_data(exceptions=()),
        output_path=output_path,
    )

    assert second.exception_row_count == 0
    workbook = load_workbook(output_path)
    assert workbook[EXCEPTION_SHEET_NAME].max_row == 1
