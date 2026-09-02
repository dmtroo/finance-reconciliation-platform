from __future__ import annotations

import argparse
import os
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from psycopg import sql

from finance_reconciliation.ingestion.database import connect

DAILY_SHEET_NAME = "Daily Summary"
EXCEPTION_SHEET_NAME = "Exceptions"

DAILY_RELATION = "mart_finance_daily"
EXCEPTION_RELATION = "mart_reconciliation_exceptions"

DEFAULT_REPORT = Path(
    "reports/exports/finance_reconciliation_report.xlsx"
)

# Export fidelity tolerance (not a reconciliation threshold): openpyxl
# round-trips numeric cells through float.
MONEY_TOLERANCE = Decimal("0.01")

# Daily Summary columns cross-checked against the mart aggregates.
COUNT_COLUMNS = (
    "exception_count",
    "unvalued_capture_count",
)
MONEY_COLUMNS = (
    "valued_capture_amount_eur",
    "reconciled_capture_amount_eur",
)

SCENARIO_ANY = "any"
SCENARIO_CLEAN = "clean"
SCENARIO_WITH_ANOMALIES = "with_anomalies"

# The frozen M5 anomaly taxonomy - every code the with_anomalies run
# must surface. Kept in sync with scripts/validate_m5_anomalies.py.
EXPECTED_EXCEPTION_CODES = frozenset(
    {
        "MISSING_CAPTURE",
        "CAPTURE_AMOUNT_MISMATCH",
        "DUPLICATE_CAPTURE",
        "INVALID_REFUND",
        "OVER_REFUND",
        "MISSING_SETTLEMENT",
        "LATE_SETTLEMENT",
        "SETTLEMENT_TOTAL_MISMATCH",
        "MISSING_BANK_RECEIPT",
        "BANK_AMOUNT_MISMATCH",
        "MISSING_LEDGER_POSTING",
        "LEDGER_AMOUNT_MISMATCH",
        "UNBALANCED_JOURNAL",
        "MISSING_FX_RATE",
        "FX_RATE_OUTLIER",
        "UNMAPPED_PRODUCT",
    }
)


class FinanceReportValidationError(RuntimeError):
    """Raised when the Excel report does not represent the current marts."""


def _sheet_rows(worksheet) -> list[tuple[object, ...]]:
    return [
        row
        for row in worksheet.iter_rows(
            values_only=True
        )
    ]


def _column_index(
    header: tuple[object, ...],
    *,
    name: str,
) -> int:
    for index, value in enumerate(header):
        if value == name:
            return index

    raise FinanceReportValidationError(
        f"Report is missing the {name!r} column"
    )


def _to_decimal(value: object) -> Decimal:
    if value is None:
        return Decimal(0)

    return Decimal(str(value))


def read_report(
    report_path: Path,
) -> dict[str, object]:
    if not report_path.exists():
        raise FinanceReportValidationError(
            f"Report file does not exist: {report_path}"
        )

    if report_path.stat().st_size == 0:
        raise FinanceReportValidationError(
            f"Report file is empty: {report_path}"
        )

    try:
        workbook = load_workbook(
            report_path,
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise FinanceReportValidationError(
            f"Report workbook could not be opened: {exc}"
        ) from exc

    missing_sheets = {
        DAILY_SHEET_NAME,
        EXCEPTION_SHEET_NAME,
    } - set(workbook.sheetnames)

    if missing_sheets:
        raise FinanceReportValidationError(
            "Report is missing sheets: "
            f"{sorted(missing_sheets)}"
        )

    daily = _sheet_rows(workbook[DAILY_SHEET_NAME])
    exceptions = _sheet_rows(
        workbook[EXCEPTION_SHEET_NAME]
    )
    workbook.close()

    if not daily or not exceptions:
        raise FinanceReportValidationError(
            "Report sheets are missing their header row"
        )

    daily_header, daily_data = daily[0], daily[1:]
    exception_header, exception_data = (
        exceptions[0],
        exceptions[1:],
    )

    exception_code_index = _column_index(
        exception_header,
        name="exception_code",
    )
    exception_codes = {
        str(row[exception_code_index])
        for row in exception_data
    }

    count_totals = {
        name: sum(
            int(row[_column_index(daily_header, name=name)] or 0)
            for row in daily_data
        )
        for name in COUNT_COLUMNS
    }

    money_totals = {
        name: sum(
            (
                _to_decimal(
                    row[
                        _column_index(
                            daily_header,
                            name=name,
                        )
                    ]
                )
                for row in daily_data
            ),
            Decimal(0),
        )
        for name in MONEY_COLUMNS
    }

    return {
        "daily_row_count": len(daily_data),
        "exception_row_count": len(exception_data),
        "exception_codes": exception_codes,
        "count_totals": count_totals,
        "money_totals": money_totals,
    }


def read_marts(
    *,
    analytics_schema: str,
) -> dict[str, object]:
    with (
        connect() as connection,
        connection.cursor() as cursor,
    ):
        cursor.execute(
            sql.SQL(
                "select count(*) from {}.{}"
            ).format(
                sql.Identifier(analytics_schema),
                sql.Identifier(DAILY_RELATION),
            )
        )
        daily_row_count = int(cursor.fetchone()[0])

        cursor.execute(
            sql.SQL(
                "select count(*) from {}.{}"
            ).format(
                sql.Identifier(analytics_schema),
                sql.Identifier(EXCEPTION_RELATION),
            )
        )
        exception_row_count = int(
            cursor.fetchone()[0]
        )

        cursor.execute(
            sql.SQL(
                """
                select
                    coalesce(sum(exception_count), 0),
                    coalesce(sum(unvalued_capture_count), 0),
                    coalesce(sum(valued_capture_amount_eur), 0),
                    coalesce(sum(reconciled_capture_amount_eur), 0)
                from {}.{}
                """
            ).format(
                sql.Identifier(analytics_schema),
                sql.Identifier(DAILY_RELATION),
            )
        )
        (
            exception_count_total,
            unvalued_capture_count_total,
            valued_amount_total,
            reconciled_amount_total,
        ) = cursor.fetchone()

        cursor.execute(
            sql.SQL(
                "select distinct exception_code from {}.{}"
            ).format(
                sql.Identifier(analytics_schema),
                sql.Identifier(EXCEPTION_RELATION),
            )
        )
        exception_codes = {
            str(row[0]) for row in cursor.fetchall()
        }

    return {
        "daily_row_count": daily_row_count,
        "exception_row_count": exception_row_count,
        "exception_codes": exception_codes,
        "count_totals": {
            "exception_count": int(exception_count_total),
            "unvalued_capture_count": int(
                unvalued_capture_count_total
            ),
        },
        "money_totals": {
            "valued_capture_amount_eur": Decimal(
                valued_amount_total
            ),
            "reconciled_capture_amount_eur": Decimal(
                reconciled_amount_total
            ),
        },
    }


def validate_finance_report(
    *,
    report_path: Path,
    analytics_schema: str,
    scenario: str = SCENARIO_ANY,
) -> None:
    report = read_report(report_path)
    marts = read_marts(
        analytics_schema=analytics_schema
    )

    if (
        report["daily_row_count"]
        != marts["daily_row_count"]
    ):
        raise FinanceReportValidationError(
            "Daily Summary row count does not match "
            f"{DAILY_RELATION}: "
            f"report={report['daily_row_count']}, "
            f"mart={marts['daily_row_count']}"
        )

    if (
        report["exception_row_count"]
        != marts["exception_row_count"]
    ):
        raise FinanceReportValidationError(
            "Exceptions row count does not match "
            f"{EXCEPTION_RELATION}: "
            f"report={report['exception_row_count']}, "
            f"mart={marts['exception_row_count']}"
        )

    for name in COUNT_COLUMNS:
        if (
            report["count_totals"][name]
            != marts["count_totals"][name]
        ):
            raise FinanceReportValidationError(
                f"Daily Summary sum({name}) does not match the mart: "
                f"report={report['count_totals'][name]}, "
                f"mart={marts['count_totals'][name]}"
            )

    for name in MONEY_COLUMNS:
        difference = abs(
            report["money_totals"][name]
            - marts["money_totals"][name]
        )
        if difference > MONEY_TOLERANCE:
            raise FinanceReportValidationError(
                f"Daily Summary sum({name}) does not match the mart: "
                f"report={report['money_totals'][name]}, "
                f"mart={marts['money_totals'][name]}, "
                f"difference={difference}"
            )

    if (
        report["exception_codes"]
        != marts["exception_codes"]
    ):
        raise FinanceReportValidationError(
            "Exceptions sheet exception_code set does not match "
            f"{EXCEPTION_RELATION}: "
            f"report={sorted(report['exception_codes'])}, "
            f"mart={sorted(marts['exception_codes'])}"
        )

    _validate_scenario(scenario, report=report)

    print(
        f"Report file: {report_path} "
        f"({report_path.stat().st_size:,} bytes)."
    )
    print(
        "Daily Summary rows: "
        f"{report['daily_row_count']:,} "
        f"(= {DAILY_RELATION})."
    )
    print(
        "Exceptions rows: "
        f"{report['exception_row_count']:,} "
        f"(= {EXCEPTION_RELATION})."
    )
    print(
        "Daily Summary totals match the mart "
        "within export tolerance."
    )
    print()
    print("Finance report validation passed.")


def _validate_scenario(
    scenario: str,
    *,
    report: dict[str, object],
) -> None:
    if scenario == SCENARIO_ANY:
        return

    exception_row_count = report["exception_row_count"]
    exception_codes = report["exception_codes"]

    if scenario == SCENARIO_CLEAN:
        if exception_row_count != 0:
            raise FinanceReportValidationError(
                "clean scenario expected 0 Exceptions rows, found "
                f"{exception_row_count}"
            )
        print(
            "Scenario clean: Exceptions sheet has 0 data rows."
        )
        return

    if scenario == SCENARIO_WITH_ANOMALIES:
        if exception_row_count == 0:
            raise FinanceReportValidationError(
                "with_anomalies scenario expected Exceptions rows, "
                "found none"
            )

        missing = sorted(
            EXPECTED_EXCEPTION_CODES - exception_codes
        )
        unexpected = sorted(
            exception_codes - EXPECTED_EXCEPTION_CODES
        )
        if missing or unexpected:
            raise FinanceReportValidationError(
                "with_anomalies scenario exception codes do not match "
                "the frozen taxonomy: "
                f"missing={missing}, unexpected={unexpected}"
            )
        print(
            "Scenario with_anomalies: all "
            f"{len(EXPECTED_EXCEPTION_CODES)} frozen exception codes "
            "present in the report."
        )
        return

    raise FinanceReportValidationError(
        f"Unknown scenario: {scenario!r}"
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Validate that the exported Finance Excel report represents "
            "the current reconciliation marts."
        )
    )

    parser.add_argument(
        "--report",
        type=Path,
        default=DEFAULT_REPORT,
    )

    parser.add_argument(
        "--analytics-schema",
        default=os.getenv("DBT_SCHEMA", "analytics_dev"),
        help="Schema that holds the reconciliation marts.",
    )

    parser.add_argument(
        "--scenario",
        choices=(
            SCENARIO_ANY,
            SCENARIO_CLEAN,
            SCENARIO_WITH_ANOMALIES,
        ),
        default=SCENARIO_ANY,
        help=(
            "Extra scenario assertion: 'clean' requires 0 Exceptions "
            "rows; 'with_anomalies' requires all 16 frozen exception "
            "codes. 'any' (default) only checks report vs marts."
        ),
    )

    args = parser.parse_args()

    validate_finance_report(
        report_path=args.report,
        analytics_schema=args.analytics_schema,
        scenario=args.scenario,
    )


if __name__ == "__main__":
    main()
